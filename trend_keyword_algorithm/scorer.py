"""SNS 트렌드 연관키워드 점수 계산 MVP."""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


TREND_SHEET = "트렌드 마스터"
KEYWORD_SHEET = "키워드 마스터"
OUTPUT_COLUMNS = [
    "trend_id", "대표 트렌드명", "후보 키워드", "역할",
    "별칭·표기", "문맥 공동언급", "시간 동조", "관계 구체성", "소비 전환성",
    "일반성 감점", "계절성 감점", "총점", "등급", "최종 판정",
]
SCORE_COLUMNS = OUTPUT_COLUMNS[4:11]
CORE_ROLES = {"대표명·별칭", "구성요소", "제품·형태"}


def normalize_header(value: object) -> str:
    """줄바꿈과 '(0~25)' 같은 점수 범위 표기를 제거한다."""
    text = "" if value is None else str(value).strip()
    text = text.replace("\n", " ")
    text = re.sub(r"\s*\(\s*\d+\s*~\s*\d+\s*\)\s*$", "", text)
    return re.sub(r"\s+", " ", text).strip()


def find_header(ws, required: Iterable[str]) -> Tuple[int, Dict[str, int]]:
    required_set = set(required)
    for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
        mapping = {
            normalize_header(value): index
            for index, value in enumerate(row)
            if normalize_header(value)
        }
        if required_set.issubset(mapping):
            return row_number, mapping
    missing = ", ".join(sorted(required_set))
    raise ValueError(f"필수 헤더 행을 찾을 수 없습니다: {missing}")


def nonempty(value: object) -> bool:
    return value is not None and str(value).strip() != ""


def score_value(value: object, column: str, row_number: int) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"키워드 마스터 {row_number}행 '{column}' 값이 숫자가 아닙니다: {value!r}")
    return value


def display_number(value: float):
    return int(value) if float(value).is_integer() else value


def calculate_grade(total: float, general_penalty: float, seasonal_penalty: float) -> str:
    if general_penalty >= 12 or seasonal_penalty >= 10 or total < 45:
        return "제외"
    if total >= 75:
        return "핵심"
    if total >= 60:
        return "연관"
    if total >= 55:
        return "소비 신호"
    return "확인중"


def calculate_final(
    total: float,
    context: float,
    timing: float,
    specificity: float,
    general_penalty: float,
    seasonal_penalty: float,
    role: str,
) -> str:
    if general_penalty >= 12 or seasonal_penalty >= 10 or total < 45:
        return "제외"
    if total >= 75 and context >= 17 and timing >= 12 and role in CORE_ROLES:
        return "핵심"
    if role == "소비 신호" and total >= 55 and specificity >= 1:
        return "소비 신호"
    if total >= 60 and specificity >= 1:
        return "연관"
    return "확인중"


def find_latest_input(workspace: Path, algorithm_dir: Path) -> Path:
    candidates = [
        path for path in workspace.rglob("*.xlsx")
        if algorithm_dir not in path.parents and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(f"워크스페이스에서 입력 .xlsx 파일을 찾지 못했습니다: {workspace}")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def read_valid_trends(workbook) -> Dict[str, str]:
    ws = workbook[TREND_SHEET]
    header_row, columns = find_header(ws, {"trend_id", "대표 트렌드명", "상태"})
    valid: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        trend_id = row[columns["trend_id"]]
        name = row[columns["대표 트렌드명"]]
        status = row[columns["상태"]]
        if nonempty(trend_id) and str(status).strip() == "검증 통과":
            valid[str(trend_id).strip()] = "" if name is None else str(name).strip()
    return valid


def build_results(workbook, valid_trends: Mapping[str, str]) -> List[dict]:
    ws = workbook[KEYWORD_SHEET]
    required = {"trend_id", "대표 트렌드", "후보 키워드", "역할", *SCORE_COLUMNS}
    header_row, columns = find_header(ws, required)
    results: List[dict] = []

    for row_number, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1
    ):
        raw_id = row[columns["trend_id"]]
        keyword = row[columns["후보 키워드"]]
        if not (nonempty(raw_id) and nonempty(keyword)):
            continue
        trend_id = str(raw_id).strip()
        if trend_id not in valid_trends:
            continue

        scores = {
            column: score_value(row[columns[column]], column, row_number)
            for column in SCORE_COLUMNS
        }
        total = (
            scores["별칭·표기"] + scores["문맥 공동언급"] + scores["시간 동조"]
            + scores["관계 구체성"] + scores["소비 전환성"]
            - scores["일반성 감점"] - scores["계절성 감점"]
        )
        role = "" if row[columns["역할"]] is None else str(row[columns["역할"]]).strip()
        result = {
            "trend_id": trend_id,
            "대표 트렌드명": valid_trends[trend_id],
            "후보 키워드": str(keyword).strip(),
            "역할": role,
            **{key: display_number(value) for key, value in scores.items()},
            "총점": display_number(total),
            "등급": calculate_grade(total, scores["일반성 감점"], scores["계절성 감점"]),
            "최종 판정": calculate_final(
                total, scores["문맥 공동언급"], scores["시간 동조"],
                scores["관계 구체성"], scores["일반성 감점"],
                scores["계절성 감점"], role,
            ),
        }
        results.append(result)
    return results


def write_csv(path: Path, rows: Sequence[Mapping[str, object]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, rows: Sequence[Mapping[str, object]]) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "result_keywords"
    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append([row[column] for column in OUTPUT_COLUMNS])

    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, column in enumerate(OUTPUT_COLUMNS, 1):
        values = [str(column), *(str(row[column]) for row in rows)]
        ws.column_dimensions[get_column_letter(index)].width = min(max(map(len, values)) + 2, 24)
    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="트렌드 연관키워드 점수 계산")
    parser.add_argument("--input", type=Path, help="입력 XLSX 경로 (생략 시 워크스페이스 최신 파일)")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    algorithm_dir = Path(__file__).resolve().parent
    workspace = algorithm_dir.parent
    input_path = args.input.resolve() if args.input else find_latest_input(workspace, algorithm_dir)
    output_dir = algorithm_dir / "output"
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        valid_trends = read_valid_trends(workbook)
        results = build_results(workbook, valid_trends)
    finally:
        workbook.close()

    csv_path = output_dir / "result_keywords.csv"
    xlsx_path = output_dir / "result_keywords.xlsx"
    write_csv(csv_path, results)
    write_xlsx(xlsx_path, results)
    print(f"입력: {input_path}")
    print(f"검증 통과 트렌드: {len(valid_trends)}건")
    print(f"결과 키워드: {len(results)}건")
    print(f"생성: {csv_path}")
    print(f"생성: {xlsx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
