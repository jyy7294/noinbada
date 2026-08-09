"""후보 수집과 점수 검토를 잇는 안전한 실행 도우미."""
from __future__ import annotations
import argparse
import shutil
import subprocess
import sys
from pathlib import Path
from openpyxl import load_workbook
from candidate_generator import generate, REVIEW_HEADERS

SCORE_START = 12; SCORE_END = 19  # M:S, zero-based

def scores_complete(review: Path) -> bool:
    wb = load_workbook(review, read_only=True, data_only=True)
    try:
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[2] and any(value is None or value == "" for value in row[SCORE_START:SCORE_END]): return False
        return True
    finally: wb.close()

def stage_scoring_workbook(workbook: Path, review: Path, staged: Path) -> None:
    """원본을 복사한 뒤 검토 시트를 키워드 마스터로 넣는다. 원본은 수정하지 않는다."""
    shutil.copy2(workbook, staged)
    staged_wb = load_workbook(staged)
    review_wb = load_workbook(review, read_only=True, data_only=True)
    try:
        if "키워드 마스터" in staged_wb.sheetnames:
            del staged_wb["키워드 마스터"]
        ws = staged_wb.create_sheet("키워드 마스터")
        for row in review_wb.active.iter_rows(values_only=True): ws.append(list(row))
        staged_wb.save(staged)
    finally:
        review_wb.close(); staged_wb.close()

def main() -> int:
    p = argparse.ArgumentParser(description="후보 생성 후 점수 검토 상태를 확인")
    p.add_argument('--workbook', type=Path, required=True); p.add_argument('--documents', type=Path)
    p.add_argument('--aliases', type=Path); p.add_argument('--output-dir', type=Path, required=True)
    p.add_argument('--review', type=Path, help='이미 점수를 입력한 candidate_review_template.xlsx 경로')
    p.add_argument('--allow-overwrite-results', action='store_true', help='기존 result_keywords 파일 덮어쓰기를 명시적으로 허용')
    a = p.parse_args()
    if a.review:
        review = a.review
    else:
        if not a.documents or not a.aliases:
            p.error('--review 없이 실행할 때는 --documents와 --aliases가 필요합니다.')
        generate(a.workbook, a.documents, a.aliases, a.output_dir)
        review = a.output_dir / 'candidate_review_template.xlsx'
    if not scores_complete(review):
        print('후보 수집 완료, 점수 검토가 필요합니다. M:S 점수를 입력한 뒤 다시 실행하세요.')
        return 2
    if not a.allow_overwrite_results:
        print('점수 입력은 완료됐지만 기존 result_keywords 보호를 위해 scorer.py 실행을 중단했습니다. 덮어쓰기를 허용하려면 --allow-overwrite-results를 사용하세요.')
        return 3
    a.output_dir.mkdir(parents=True, exist_ok=True)
    staged = a.output_dir / 'staged_scoring_input.xlsx'
    stage_scoring_workbook(a.workbook, review, staged)
    scorer = Path(__file__).resolve().parent / 'scorer.py'
    return subprocess.run([sys.executable, str(scorer), '--input', str(staged)]).returncode

if __name__ == '__main__': raise SystemExit(main())
