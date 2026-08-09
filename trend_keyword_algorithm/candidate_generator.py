"""사람이 제공한 문서 CSV에서 검토용 트렌드 키워드 후보를 만든다.

외부 서비스에 접속하지 않으며, 입력 CSV의 내용만 처리한다.
"""
from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from scorer import find_header, nonempty


REVIEW_HEADERS = [
    "trend_id", "대표 트렌드", "후보 키워드", "역할", "정규화 키워드", "별칭 그룹",
    "출처 유형", "근거 URL/ID", "근거 날짜", "인용 문맥/메모", "동시 언급 수",
    "확산월 시차(개월)", "별칭·표기\n(0~25)", "문맥 공동언급\n(0~25)",
    "시간 동조\n(0~20)", "관계 구체성\n(0~15)", "소비 전환성\n(0~15)",
    "일반성 감점\n(0~20)", "계절성 감점\n(0~15)", "총점", "등급", "최종 판정",
]
POOL_HEADERS = [
    "trend_id", "대표 트렌드명", "candidate_keyword", "normalized_keyword", "role_guess",
    "co_mention_count", "source_count", "first_seen", "peak_month", "peak_month_delta",
    "evidence_urls", "evidence_memo", "review_status",
]
EVIDENCE_HEADERS = [
    "trend_id", "normalized_keyword", "candidate_keyword", "platform", "published_at", "url", "text_excerpt",
]
BEHAVIOR_WORDS = ("레시피", "만들기", "챌린지", "안무", "품절", "구매처", "오픈런", "맛집", "후기", "먹방")
CONTEXT_TERMS = ("뉴진스", "newjeans", "브랜드", "프랜차이즈", "밈", "아티스트")


def normalize_keyword(value: str) -> str:
    value = value.strip().lstrip("#")
    value = re.sub(r"\s+", " ", value)
    return value.casefold().replace(" ", "")


def clean_text(text: str) -> str:
    text = re.sub(r"https?://\S+|www\.\S+|\S+@\S+", " ", text)
    text = re.sub(r"@[A-Za-z0-9_가-힣]+", " ", text)
    return re.sub(r"[^0-9A-Za-z가-힣#\s]", " ", text)


def read_csv(path: Path, required: Sequence[str]) -> List[dict]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None or not set(required).issubset(reader.fieldnames):
            raise ValueError(f"{path}의 헤더는 다음을 포함해야 합니다: {', '.join(required)}")
        return [row for row in reader if any(nonempty(value) for value in row.values())]


def read_valid_trends(workbook_path: Path) -> Dict[str, dict]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb["트렌드 마스터"]
        header_row, columns = find_header(ws, {"trend_id", "대표 트렌드명", "상태"})
        month_column = columns.get("최초 확산월 가설", columns.get("관찰 시작월"))
        trends = {}
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            trend_id, status = row[columns["trend_id"]], row[columns["상태"]]
            if nonempty(trend_id) and str(status).strip() == "검증 통과":
                month = row[month_column] if month_column is not None else None
                trends[str(trend_id).strip()] = {
                    "name": str(row[columns["대표 트렌드명"]]).strip(), "month": month,
                }
        return trends
    finally:
        wb.close()


def role_guess(keyword: str, aliases: Iterable[str]) -> str:
    lowered = keyword.casefold()
    if normalize_keyword(keyword) in {normalize_keyword(alias) for alias in aliases}:
        return "대표명·별칭"
    if any(word in keyword for word in BEHAVIOR_WORDS):
        return "소비 신호"
    if any(term in lowered for term in CONTEXT_TERMS):
        return "확산 맥락"
    return "확인중"


def extract_candidates(text: str, aliases: Sequence[str]) -> List[str]:
    """해시태그, 등록 별칭, 행동 결합 표현, 인용된 표현을 정규식으로 추출한다."""
    cleaned = clean_text(text)
    found = re.findall(r"#([A-Za-z가-힣0-9]{2,})", cleaned)
    for alias in aliases:
        if alias and alias.casefold() in cleaned.casefold():
            found.append(alias)
    action_pattern = "|".join(BEHAVIOR_WORDS)
    for alias in aliases:
        # 등록된 표기 뒤에 붙은 행동 표현을 보존한다. 조사(를/은 등)는 후보명에 넣지 않는다.
        alias_pattern = re.escape(alias).replace(r"\ ", r"\s+")
        found.extend(re.findall(rf"({alias_pattern}\s+(?:{action_pattern}))(?:[은는이가을를의와과도만]{{0,2}})?", cleaned, flags=re.IGNORECASE))
        found.extend(re.findall(rf"({alias_pattern}(?:\s+[A-Za-z가-힣0-9]+){{0,2}}\s+(?:{action_pattern}))(?:[은는이가을를의와과도만]{{0,2}})?", cleaned, flags=re.IGNORECASE))
    # 확산 맥락으로 자주 나타나는 조사 결합 표현도 보존한다.
    found.extend(re.findall(r"[A-Za-z가-힣0-9]+의\s+[A-Za-z가-힣0-9]+", cleaned))
    return [re.sub(r"\s+", " ", item).strip() for item in found if len(normalize_keyword(item)) >= 2]


def load_stopwords() -> set:
    path = Path(__file__).resolve().parent / "data" / "stopwords_ko.txt"
    return {normalize_keyword(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()}


def month_delta(first: object, baseline: object):
    if not first or not baseline:
        return ""
    try:
        first_date = datetime.strptime(str(first)[:10], "%Y-%m-%d")
        if isinstance(baseline, datetime):
            baseline_date = baseline
        else:
            baseline_date = datetime.fromisoformat(str(baseline))
        return (first_date.year - baseline_date.year) * 12 + first_date.month - baseline_date.month
    except (TypeError, ValueError):
        return ""


def write_csv(path: Path, headers: Sequence[str], rows: Sequence[Mapping[str, object]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def write_review_xlsx(path: Path, rows: Sequence[Mapping[str, object]]) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "키워드 마스터"
    ws.append(REVIEW_HEADERS)
    for item in rows:
        ws.append([
            item["trend_id"], item["대표 트렌드명"], item["candidate_keyword"], item["role_guess"],
            item["normalized_keyword"], "", "사람 입력", item["evidence_urls"], item["first_seen"],
            item["evidence_memo"], item["co_mention_count"], item["peak_month_delta"],
            "", "", "", "", "", "", "", "", "", "",
        ])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78"); cell.font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = 18
    wb.save(path)


def generate(workbook: Path, documents: Path, aliases_path: Path, output_dir: Path) -> List[dict]:
    trends = read_valid_trends(workbook)
    documents_rows = read_csv(documents, ("trend_id", "platform", "published_at", "text", "url"))
    alias_rows = read_csv(aliases_path, ("trend_id", "alias", "normalized_alias"))
    aliases = defaultdict(list)
    for row in alias_rows:
        if row["trend_id"] in trends and nonempty(row["alias"]): aliases[row["trend_id"]].append(row["alias"].strip())
    for trend_id, trend in trends.items(): aliases[trend_id].append(trend["name"])

    stopwords = load_stopwords()
    grouped = defaultdict(list)
    seen_documents = set()
    for doc in documents_rows:
        trend_id = doc["trend_id"].strip()
        if trend_id not in trends or not nonempty(doc["text"]): continue
        text, url = doc["text"].strip(), doc["url"].strip()
        if not any(alias.casefold() in text.casefold() for alias in aliases[trend_id]): continue
        key = (normalize_keyword(text), url)
        if key in seen_documents: continue
        seen_documents.add(key)
        for candidate in extract_candidates(text, aliases[trend_id]):
            normalized = normalize_keyword(candidate)
            if normalized in stopwords: continue
            grouped[(trend_id, normalized)].append((candidate, doc))

    pools, evidence = [], []
    for (trend_id, normalized), mentions in sorted(grouped.items()):
        unique_docs = {(item[1]["url"], normalize_keyword(item[1]["text"])): item for item in mentions}
        representative = next((item[0] for item in mentions if not item[0].startswith("#")), mentions[0][0].lstrip("#"))
        alias_set = aliases[trend_id]
        is_alias = normalize_keyword(representative) in {normalize_keyword(alias) for alias in alias_set}
        if len(unique_docs) < 2 and not is_alias: continue
        docs = list(unique_docs.values())
        dates = sorted(item[1]["published_at"] for item in docs if item[1]["published_at"])
        platforms = {item[1]["platform"] for item in docs if item[1]["platform"]}
        urls = list(dict.fromkeys(item[1]["url"] for item in docs if item[1]["url"]))[:3]
        first_seen = dates[0] if dates else ""
        memo = clean_text(docs[0][1]["text"]).strip()[:180]
        row = {
            "trend_id": trend_id, "대표 트렌드명": trends[trend_id]["name"],
            "candidate_keyword": representative.lstrip("#"), "normalized_keyword": normalized,
            "role_guess": role_guess(representative, alias_set), "co_mention_count": len(unique_docs),
            "source_count": len(platforms), "first_seen": first_seen,
            "peak_month": first_seen[:7] if first_seen else "",
            "peak_month_delta": month_delta(first_seen, trends[trend_id]["month"]),
            "evidence_urls": " | ".join(urls), "evidence_memo": memo, "review_status": "확인중",
        }
        pools.append(row)
        for _, doc in docs:
            evidence.append({"trend_id": trend_id, "normalized_keyword": normalized,
                "candidate_keyword": row["candidate_keyword"], "platform": doc["platform"],
                "published_at": doc["published_at"], "url": doc["url"], "text_excerpt": clean_text(doc["text"]).strip()[:180]})

    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(output_dir / "candidate_pool.csv", POOL_HEADERS, pools)
    write_csv(output_dir / "candidate_evidence.csv", EVIDENCE_HEADERS, evidence)
    write_review_xlsx(output_dir / "candidate_review_template.xlsx", pools)
    return pools


def main() -> int:
    parser = argparse.ArgumentParser(description="오프라인 SNS 문서 기반 후보 키워드 생성")
    parser.add_argument("--workbook", type=Path, required=True); parser.add_argument("--documents", type=Path, required=True)
    parser.add_argument("--aliases", type=Path, required=True); parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args(); rows = generate(args.workbook, args.documents, args.aliases, args.output_dir)
    print(f"후보 {len(rows)}건 생성: {args.output_dir}"); return 0


if __name__ == "__main__": raise SystemExit(main())
